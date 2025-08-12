"""
Advanced features module for Excel file manipulation.
Includes password support, empty cell validation, conditional formatting,
formula manipulation and chart support.
"""

import io
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from msoffcrypto import OfficeFile
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


def read_protected_excel(file_path: str, password: str) -> pd.DataFrame:
    """
    Reads a password-protected Excel file.

    Args:
        file_path (str): Path to the Excel file
        password (str): File password

    Returns:
        pd.DataFrame: DataFrame with file data

    Raises:
        ValueError: If password is incorrect
        FileNotFoundError: If file doesn't exist
    """
    try:
        with open(file_path, "rb") as file:
            office_file = OfficeFile(file)
            office_file.load_key(password=password)

            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)

            return pd.read_excel(decrypted)
    except Exception as e:
        raise ValueError(f"Error reading protected file: {str(e)}")


def validate_empty_cells(
    df: pd.DataFrame, columns: Optional[List[str]] = None, threshold: float = 0.1
) -> Dict[str, Any]:
    """
    Validates empty cells in a DataFrame.

    Args:
        df (pd.DataFrame): DataFrame to be validated
        columns (List[str], optional): List of columns to validate. If None, validates all.
        threshold (float): Maximum percentage of empty cells allowed (0-1)

    Returns:
        Dict[str, Any]: Dictionary with validation results
    """  # noqa: E501
    if columns is None:
        columns = df.columns.tolist()

    results = {
        "total_cells": len(df) * len(columns),
        "empty_cells": {},
        "columns_above_threshold": [],
    }

    for col in columns:
        empty_count = df[col].isna().sum()
        empty_percent = empty_count / len(df)
        results["empty_cells"][col] = {"count": empty_count, "percent": empty_percent}

        if empty_percent > threshold:
            results["columns_above_threshold"].append(col)

    return results


def apply_conditional_formatting(file_path: str, rules: List[Dict[str, Any]]) -> None:
    """
    Applies conditional formatting to an Excel file.

    Args:
        file_path (str): Path to the Excel file
        rules (List[Dict[str, Any]]): List of formatting rules
            Each rule must contain:
            - 'range': cell range (e.g., 'A1:B10')
            - 'type': formatting type ('cellIs', 'containsText', etc.)
            - 'operator': operator ('greaterThan', 'lessThan', etc.)
            - 'formula': formula or value for comparison
            - 'format': style dictionary (e.g., {'fill': 'FF0000'})
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for rule in rules:
        cell_range = rule["range"]
        fmt = rule["format"]

        for row in ws[cell_range]:
            for cell in row:
                if rule["type"] == "cellIs":
                    try:
                        # Convert cell value to number if possible
                        cell_value = float(cell.value) if cell.value is not None else 0
                        formula_value = float(rule["formula"])

                        # Map operators to comparison functions
                        operators = {
                            ">": lambda x, y: x > y,
                            "<": lambda x, y: x < y,
                            ">=": lambda x, y: x >= y,
                            "<=": lambda x, y: x <= y,
                            "==": lambda x, y: x == y,
                            "!=": lambda x, y: x != y,
                        }

                        if operators[rule["operator"]](cell_value, formula_value):
                            if "fill" in fmt:
                                # Add 'FF' at the beginning for full opacity
                                fill_color = f"FF{fmt['fill']}"
                                cell.fill = PatternFill(
                                    start_color=fill_color,
                                    end_color=fill_color,
                                    fill_type="solid",
                                )
                            if "font" in fmt:
                                cell.font = Font(**fmt["font"])
                    except (ValueError, TypeError):
                        # Ignore cells that cannot be converted to number
                        continue

    wb.save(file_path)


def extract_formulas(file_path: str) -> Dict[str, List[Dict[str, str]]]:
    """
    Extracts formulas from an Excel file.

    Args:
        file_path (str): Path to the Excel file

    Returns:
        Dict[str, List[Dict[str, str]]]: Dictionary with formulas per sheet
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    formulas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    sheet_formulas.append(
                        {"cell": cell.coordinate, "formula": cell.value}
                    )

        formulas[sheet_name] = sheet_formulas

    return formulas


def add_chart(
    file_path: str,
    chart_type: str,
    data_range: str,
    title: str,
    output_file: Optional[str] = None,
) -> None:
    """
    Adds a chart to an Excel file.

    Args:
        file_path (str): Path to the Excel file
        chart_type (str): Chart type ('bar', 'line', 'pie')
        data_range (str): Data range (e.g., 'A1:B10')
        title (str): Chart title
        output_file (str, optional): Path to save the modified file
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Create chart based on type
    if chart_type == "bar":
        chart = BarChart()
    # Add other chart types here

    # Define data including sheet name
    data_range_with_sheet = f"{ws.title}!{data_range}"
    data = Reference(ws, range_string=data_range_with_sheet)
    chart.add_data(data, titles_from_data=True)

    # Configure chart
    chart.title = title
    chart.style = 13

    # Add chart to sheet
    ws.add_chart(chart, "E5")

    # Save file
    output_path = output_file or file_path
    wb.save(output_path)


def protect_excel(
    file_path: str, password: str, output_file: Optional[str] = None
) -> None:
    """
    Protects an Excel file with a password.

    Args:
        file_path (str): Path to the Excel file
        password (str): Password to protect the file
        output_file (str, optional): Path to save the protected file
    """
    wb = openpyxl.load_workbook(file_path)

    # Protect all worksheets
    for ws in wb.worksheets:
        ws.protection.set_password(password)

    # Save file
    output_path = output_file or file_path
    wb.save(output_path)
