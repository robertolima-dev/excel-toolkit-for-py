"""
Testes para as funcionalidades avançadas do excel-toolkit-for-py
"""

import os
import pytest
import pandas as pd
import openpyxl
from excel_toolkit_for_py.advanced_features import (
    read_protected_excel,
    validate_empty_cells,
    apply_conditional_formatting,
    extract_formulas,
    add_chart,
    protect_excel
)

# Dados de teste
TEST_DATA = {
    'Nome': ['Alice', 'Bob', 'Carlos', None, 'Eva'],
    'Idade': [25, 30, None, 40, 35],
    'Cidade': ['SP', 'RJ', 'BH', 'POA', None]
}

def create_test_excel():
    """Cria um arquivo Excel de teste"""
    df = pd.DataFrame(TEST_DATA)
    df.to_excel('test.xlsx', index=False)
    
    # Garante que o arquivo foi criado e fechado
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.active
    
    # Verifica se os dados foram salvos corretamente
    assert ws['B5'].value == 40, "Dados não foram salvos corretamente"
    wb.close()
    
    return 'test.xlsx'

def test_validate_empty_cells():
    """Testa a validação de células vazias"""
    df = pd.DataFrame(TEST_DATA)
    result = validate_empty_cells(df)
    
    assert result['total_cells'] == 15  # 5 linhas * 3 colunas
    assert 'Nome' in result['empty_cells']
    assert 'Idade' in result['empty_cells']
    assert 'Cidade' in result['empty_cells']
    assert len(result['columns_above_threshold']) > 0

def test_apply_conditional_formatting():
    """Testa a aplicação de formatação condicional"""
    file_path = create_test_excel()
    
    rules = [{
        'range': 'B2:B6',  # Apenas a coluna de idade
        'type': 'cellIs',
        'operator': '>',
        'formula': '30',
        'format': {
            'fill': 'FF0000',
            'font': {'bold': True}
        }
    }]
    
    apply_conditional_formatting(file_path, rules)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Verifica se a formatação foi aplicada na célula com valor maior que 30
    assert ws['B5'].value == 40
    assert ws['B5'].fill.start_color.rgb == 'FFFF0000'  # FF no início para opacidade
    assert ws['B5'].font.bold is True
    
    wb.close()
    os.remove(file_path)

def test_extract_formulas():
    """Testa a extração de fórmulas"""
    file_path = create_test_excel()
    
    # Adiciona uma fórmula
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws['D2'] = '=SUM(B2:B6)'
    wb.save(file_path)
    
    formulas = extract_formulas(file_path)
    
    assert 'Sheet1' in formulas
    assert len(formulas['Sheet1']) == 1
    assert formulas['Sheet1'][0]['cell'] == 'D2'
    assert formulas['Sheet1'][0]['formula'] == '=SUM(B2:B6)'
    
    os.remove(file_path)

def test_add_chart():
    """Testa a adição de gráficos"""
    file_path = create_test_excel()
    output_file = 'test_with_chart.xlsx'
    
    add_chart(
        file_path=file_path,
        chart_type='bar',
        data_range='A1:B6',
        title='Test Chart',
        output_file=output_file
    )
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # Verifica se o gráfico foi adicionado
    assert len(ws._charts) > 0
    
    os.remove(file_path)
    os.remove(output_file)

def test_protect_excel():
    """Testa a proteção de arquivos Excel"""
    file_path = create_test_excel()
    output_file = 'test_protected.xlsx'
    password = 'test123'
    
    protect_excel(file_path, password, output_file)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # Verifica se a proteção foi aplicada
    assert ws.protection.sheet is True
    
    os.remove(file_path)
    os.remove(output_file)

@pytest.mark.skip(reason="Necessita de arquivo Excel protegido para teste")
def test_read_protected_excel():
    """Testa a leitura de arquivos Excel protegidos"""
    # Este teste requer um arquivo Excel protegido para funcionar
    file_path = 'protected_test.xlsx'
    password = 'test123'
    
    df = read_protected_excel(file_path, password)
    assert isinstance(df, pd.DataFrame)
    assert not df.empty 